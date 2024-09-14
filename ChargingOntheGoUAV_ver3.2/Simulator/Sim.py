import math
import random
from Simulator.MIP import MIP
import ctypes
import time
from openpyxl import Workbook
from datetime import datetime
import numpy as np
from matplotlib import pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.ticker import FormatStrFormatter
import seaborn as sns
import pandas as pd

class Sim():
    def __init__(self, M, uavList, mcsList, numEVs, result_list, numUAVs, MIPFlag = True, PolicyFlag = True, PolicyDPCompare = False, MCSFlag = False):
        

        self.M = M
        self.data_Graph_Export_DP_ServiceFee_FlightCost = [] # 0: # of EVs, 1: Unit flight cost, 2: Service Fee, 3: Depot return rate, 4: Get charged rate
        self.data_Graph_Export_DP_Arm_NumEVs = [] # 0: Arm, 1: # of EVs, 2: Proceed rate, 3: Depot return rate, 4: Get charged rate, 5: AVg. DP Obj. value
        self.data_Graph_Export_DP_UAVCapacity = [] # 0: UAV Capacity, 1: proceed rate_DP, 2: return rate_DP, 3: charged rate_DP, 4: obj. value_DP, 5: proceed rate_MIP, 6: return rate_MIP, 7: charged rate_MIP, 8: obj. value_MIP
        self.uavList = uavList
        self.mcsList = mcsList
        self.numEVs = numEVs
        
        
       
        if(PolicyFlag == False):
            # Run DP algorithm
            for uav in self.uavList:
                start = time.time()
                uav.solution_DP = uav.dp.RunDP()
                end = time.time()
                self.CPUTime_DP = end - start
                print(f"{end - start:.5f} sec")
                self.objValue_DP = uav.dp.ObjectiveValue(uav.solution_DP)
                print(self.objValue_DP)
                self.feasible_DP = uav.dp.DPVerification(uav.solution_DP)
                
                
                if(self.feasible_DP == False):
                    ctypes.windll.user32.MessageBoxW(0, "DP is infeasible", "Infeasible", 1)
                    quit()
            
            
        
        else:
            # Run optimal policy
            for iter, uav in enumerate(self.uavList):
                start = time.time()
                (uav.solution_policy, uav.remainingEnergyList) = uav.dp.RunOptimalPolicy_New()  # calculate f function from scratch
                end = time.time()
                uav.CPUTime_Policy = end - start
                print(f"{end - start:.5f} sec")
                uav.objValue_Policy = uav.dp.ObjectiveValue(uav.solution_policy)
                print(uav.objValue_Policy)
                
                uav.feasible_DP = uav.dp.DPVerification(uav.solution_policy)    
                if(uav.feasible_DP == False):
                    ctypes.windll.user32.MessageBoxW(0, "DP is infeasible", "Infeasible", 1)
                    quit()
                
                uav.GenerateSchedule()
        
            
                if (MCSFlag):
            
                    # Check whether the capacity exceeds
                    self.mcsList[iter].CheckCapacity()
                    
                    # Generate schedule
                    self.mcsList[iter].GenerateSchedule()
            
            
            
                if (PolicyDPCompare == True):
                    # Run DP algorithm
                    start = time.time()
                    uav.solution_DP = uav.dp.RunDP()
                    end = time.time()
                    uav.CPUTime_DP = end - start
                    print(f"{end - start:.5f} sec")
                    uav.objValue_DP = uav.dp.ObjectiveValue(uav.solution_DP)
                    print(self.objValue_DP)
                    uav.feasible_DP = uav.dp.DPVerification(uav.solution_DP)
                    
                    
                    if(uav.feasible_DP == False):
                        ctypes.windll.user32.MessageBoxW(0, "DP is infeasible", "Infeasible", 1)
                        quit()
            
            
        
        if (MCSFlag):
            self.Evaluation_UAV_MCS(self.uavList, self.mcsList, result_list, numUAVs)
        
        
        if(MIPFlag):
            # Define MIP
            for uav in self.uavList:
                mip = MIP(uav, M)
                
                
                # Run MIP model
                start = time.time()
                (uav.x1_MIP, uav.x2_MIP, uav.y_MIP, uav.z_MIP, self.realized_q_MIP, uav.objValue_MIP, uav.feasible_MIP) = mip.RunMIP()
                
                
                end = time.time()
                uav.CPUTime_MIP = end - start
                print(f"{end - start:.5f} sec")
                
                if uav.feasible_MIP == True:
                    uav.GenerateSchedule_MIP() 
                    self.Evaluation_DP_MIP(self.uavList, result_list, numUAVs)
                
                
                if(uav.feasible_MIP):
                    print(uav.objValue_MIP)
                    # Alarm for the case where obj. value of DP is larger than that of MIP
                    # if(self.objValue_Policy > self.objValue_MIP):
                        # ctypes.windll.user32.MessageBoxW(0, "Obj. value of DP is larger than that of MIP", "Alarm", 1)
                        # quit()
                else:
                    print("MIP is infeasible")
        
    def Evaluation_DP_MIP(self, uavList, result_list, numUAVs):
        # 고객에게 서비스 완료 한 시각
        # UAV: destination 도착하는시각
        '''
        totalServiceCompletionTime_uav_DP = 0.0
        totalServiceCompletionTime_uav_MIP = 0.0
        for uav in uavList:
            for schedule in uav.scheduleList:
                totalServiceCompletionTime_uav_DP += schedule.destination_arrivalTime
            for schedule_MIP in uav.scheduleList_MIP:
                totalServiceCompletionTime_uav_MIP += schedule_MIP.destination_arrivalTime
                
        
                
        avgServiceCompletionTime_uav_DP = totalServiceCompletionTime_uav_DP / self.numEVs
        avgServiceCompletionTime_uav_MIP = totalServiceCompletionTime_uav_MIP / self.numEVs
        
        print('average service completion time (DP): ' + str(avgServiceCompletionTime_uav_DP))
        print('average service completion time (MIP): ' + str(avgServiceCompletionTime_uav_MIP))
        '''
        for uav in uavList:
            totalCnt_x2 = 0
            for i in range(len(uav.x2_MIP)):
                if round(uav.x2_MIP[i], 0) == 1:
                    totalCnt_x2 += 1
        
        
        # Result = {'numEVs': self.numEVs, 'numUAVs' : numUAVs, 'avgCompletionTime_DP': avgServiceCompletionTime_uav_DP, 'avgCompletionTime_MIP': avgServiceCompletionTime_uav_MIP, 'numWorstCase' : totalCnt_x2}
        Result = {'numWorstCase' : totalCnt_x2}
        result_list.append(Result)    
        
    def Evaluation_UAV_MCS(self, uavList, mcsList, result_list, numUAVs):
        # 고객에게 서비스 완료 한 시각
        # UAV: destination 도착하는시각
        # MCS: rendezvous 에서 MCS가 출발하는시각
        
        totalServiceCompletionTime_uav = 0.0
        totalServiceCompletionTime_mcs = 0.0
        for uav in uavList:
            for schedule in uav.scheduleList:
                totalServiceCompletionTime_uav += schedule.destination_arrivalTime
                
        for mcs in mcsList:
            for schedule in mcs.scheduleList:
                totalServiceCompletionTime_mcs += schedule.rendezvous_departureTime
                
        avgServiceCompletionTime_uav = totalServiceCompletionTime_uav / self.numEVs
        avgServiceCompletionTime_mcs = totalServiceCompletionTime_mcs / self.numEVs
        
        print('average service completion time (UAV): ' + str(avgServiceCompletionTime_uav))
        print('average service completion time (MCS): ' + str(avgServiceCompletionTime_mcs))
        
        Result = {'numEVs': self.numEVs, 'numUAVs' : numUAVs, 'numMCSs': numUAVs, 'avgCompletionTime_UAV': avgServiceCompletionTime_uav, 'avgCompletionTime_MCS': avgServiceCompletionTime_mcs}
        result_list.append(Result)
    
    def FindDistance(self):
        shortest = self.M
        longest = -999999
        
        for i in range(len(self.evList)):
            dist = self.GetKMDistance((self.depotLoc[0], self.depotLoc[1]), (self.evList[i].rendezNode.x, self.evList[i].rendezNode.y)) 
            
            if dist >= longest:
                longest = dist
                
            if dist <= shortest:
                shortest = dist        
        
        return (longest, shortest)
    
    
    def Graph_UAV_MCS(self, resultList):
            
        plt.rcParams['font.family'] = 'Times New Roman'
        plt.rc('font', size=10)
        
        
        plt.figure(1)
        
        
        df = pd.DataFrame.from_dict(resultList)
        df = df.groupby(['numEVs', 'numUAVs', 'numMCSs'], as_index=False).mean()
        # df_UAV_1 = df.loc[(df.numUAVs == 1) & (df.numMCSs == 1), ]
        # df_UAV_2 = df.loc[(df.numUAVs == 2) & (df.numMCSs == 2), ]
        # df_UAV_3 = df.loc[(df.numUAVs == 3) & (df.numMCSs == 3), ]
        # df_UAV_4 = df.loc[(df.numUAVs == 4) & (df.numMCSs == 4), ]
        # df_UAV_5 = df.loc[(df.numUAVs == 5) & (df.numMCSs == 5), ]
        
        numEVs = list(df['numEVs'].drop_duplicates())
        avgCompletionTime_UAV_1 = list(df.loc[(df.numUAVs == 1) & (df.numMCSs == 1), ]['avgCompletionTime_UAV'])
        avgCompletionTime_MCS_1 = list(df.loc[(df.numUAVs == 1) & (df.numMCSs == 1), ]['avgCompletionTime_MCS'])
        
        avgCompletionTime_UAV_2 = list(df.loc[(df.numUAVs == 2) & (df.numMCSs == 2), ]['avgCompletionTime_UAV'])
        avgCompletionTime_MCS_2 = list(df.loc[(df.numUAVs == 2) & (df.numMCSs == 2), ]['avgCompletionTime_MCS'])
        
        avgCompletionTime_UAV_3 = list(df.loc[(df.numUAVs == 3) & (df.numMCSs == 3), ]['avgCompletionTime_UAV'])
        avgCompletionTime_MCS_3 = list(df.loc[(df.numUAVs == 3) & (df.numMCSs == 3), ]['avgCompletionTime_MCS'])
        
        avgCompletionTime_UAV_4 = list(df.loc[(df.numUAVs == 4) & (df.numMCSs == 4), ]['avgCompletionTime_UAV'])
        avgCompletionTime_MCS_4 = list(df.loc[(df.numUAVs == 4) & (df.numMCSs == 4), ]['avgCompletionTime_MCS'])
        
        avgCompletionTime_UAV_5 = list(df.loc[(df.numUAVs == 5) & (df.numMCSs == 5), ]['avgCompletionTime_UAV'])
        avgCompletionTime_MCS_5 = list(df.loc[(df.numUAVs == 5) & (df.numMCSs == 5), ]['avgCompletionTime_MCS'])
        
        avgCompletionTime_UAV_6 = list(df.loc[(df.numUAVs == 6) & (df.numMCSs == 6), ]['avgCompletionTime_UAV'])
        avgCompletionTime_MCS_6 = list(df.loc[(df.numUAVs == 6) & (df.numMCSs == 6), ]['avgCompletionTime_MCS'])
        
        ##
        ax = plt.subplot(231)
        
        plt.plot(numEVs, avgCompletionTime_UAV_1, '.--', c='blue', label='avg. service completion time (UAV)')
        plt.plot(numEVs, avgCompletionTime_MCS_1, '.--', c='black', label='avg. service completion time (WMCS)')
        ax.set_title('#UAVs = #WMCSs = 1')
        
        
        
        ax.set_xticks(numEVs)
        ax.set_yticks([400, 800, 1200, 1600, 2000, 2400, 2800, 3200, 3600])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Number of EVs')
        plt.ylabel('Time (sec.)')
        
        
        
        
        
        
        ##
        ax = plt.subplot(232)
        plt.plot(numEVs, avgCompletionTime_UAV_2, '.--', c='blue', label='avg. service completion time (UAV)')
        plt.plot(numEVs, avgCompletionTime_MCS_2, '.--', c='black', label='avg. service completion time (WMCS)')
        ax.set_title('#UAVs = #WMCSs = 2')
        
        
        
        ax.set_xticks(numEVs)
        ax.set_yticks([400, 800, 1200, 1600, 2000, 2400, 2800, 3200, 3600])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Number of EVs')
        plt.ylabel('Time (sec.)')
        
        
        
        
        ##
        ax = plt.subplot(233)
        plt.plot(numEVs, avgCompletionTime_UAV_3, '.--', c='blue', label='avg. service completion time (UAV)')
        plt.plot(numEVs, avgCompletionTime_MCS_3, '.--', c='black', label='avg. service completion time (WMCS)')
        ax.set_title('#UAVs = #WMCSs = 3')
        
        
        
        ax.set_xticks(numEVs)
        ax.set_yticks([400, 800, 1200, 1600, 2000, 2400, 2800, 3200, 3600])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Number of EVs')
        plt.ylabel('Time (sec.)')
        
        
        ##
        ax = plt.subplot(234)
        plt.plot(numEVs, avgCompletionTime_UAV_4, '.--', c='blue', label='avg. service completion time (UAV)')
        plt.plot(numEVs, avgCompletionTime_MCS_4, '.--', c='black', label='avg. service completion time (WMCS)')
        ax.set_title('#UAVs = #WMCSs = 4')
        
        
        
        ax.set_xticks(numEVs)
        ax.set_yticks([400, 800, 1200, 1600, 2000, 2400, 2800, 3200, 3600])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Number of EVs')
        plt.ylabel('Time (sec.)')
        
        
        ##
        ax = plt.subplot(235)
        plt.plot(numEVs, avgCompletionTime_UAV_5, '.--', c='blue', label='avg. service completion time (UAV)')
        plt.plot(numEVs, avgCompletionTime_MCS_5, '.--', c='black', label='avg. service completion time (WMCS)')
        ax.set_title('#UAVs = #WMCSs = 5')
        
        
        
        ax.set_xticks(numEVs)
        ax.set_yticks([400, 800, 1200, 1600, 2000, 2400, 2800, 3200, 3600])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Number of EVs')
        plt.ylabel('Time (sec.)')
        
        
        ##
        ax = plt.subplot(236)
        plt.plot(numEVs, avgCompletionTime_UAV_6, '.--', c='blue', label='avg. service completion time (UAV)')
        plt.plot(numEVs, avgCompletionTime_MCS_6, '.--', c='black', label='avg. service completion time (WMCS)')
        ax.set_title('#UAVs = #WMCSs = 6')
        
        
        
        ax.set_xticks(numEVs)
        ax.set_yticks([400, 800, 1200, 1600, 2000, 2400, 2800, 3200, 3600])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Number of EVs')
        plt.ylabel('Time (sec.)')
        
        
        
        
        
        plt.legend(loc='best')
        # plt.legend(loc='best', bbox_to_anchor=(1, 1))
        plt.tight_layout()
        plt.show()   




        
    def Graph_DP_ServiceFee_FlightCost(self, armList, cost_chargeFeeList, cost_flightList, depot_return_rate_avg, charged_rate_avg, obj_DP_avg):
            
        plt.rcParams['font.family'] = 'Times New Roman'
        plt.rc('font', size=10)
        
        
        plt.figure(1)
        bar_width = 0.25
        
        ##
        ax = plt.subplot(231)
        avgReturnRate = []
        avgChargedRate = []
        avgProceedRate = []
        for k in range(len(cost_chargeFeeList)):
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            cnt = 0
            for i in range(len(self.data_Graph_Export_DP_ServiceFee_FlightCost[0])):    
                if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][1] == cost_flightList[0]):    
                    if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][2] == cost_chargeFeeList[k]):
                        sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][3]
                        sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][4]
                        cnt = cnt + 1
            
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            avgProceedRate.append(1 - (avgReturnRate[-1] + avgChargedRate[-1]))
        
        
        
        plt.plot(cost_chargeFeeList, avgReturnRate, '.--', c='blue', label='depot return rate')
        plt.plot(cost_chargeFeeList, avgChargedRate, '.--', c='black', label='getting charged via an EV rate')
        
        
        
        
        ax.set_xticks(cost_chargeFeeList)
        ax.set_yticks([0.0, 0.03, 0.06, 0.09, 0.12, 0.15, 0.18])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Service fee \n Unit flight cost = ' + str(cost_flightList[0]))
        plt.ylabel('Rate')
        
        
        
        
        
        
        ##
        ax = plt.subplot(232)
        avgReturnRate = []
        avgChargedRate = []
        avgProceedRate = []
        for k in range(len(cost_chargeFeeList)):
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            cnt = 0
            for i in range(len(self.data_Graph_Export_DP_ServiceFee_FlightCost[0])):    
                if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][1] == cost_flightList[1]):    
                    if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][2] == cost_chargeFeeList[k]):
                        sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][3]
                        sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][4]
                        cnt = cnt + 1
            
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            avgProceedRate.append(1 - (avgReturnRate[-1] + avgChargedRate[-1]))
        
        
        
        plt.plot(cost_chargeFeeList, avgReturnRate, '.--', c='blue', label='depot return rate')
        plt.plot(cost_chargeFeeList, avgChargedRate, '.--', c='black', label='getting charged via an EV rate')
        
        
        
        ax.set_xticks(cost_chargeFeeList)
        ax.set_yticks([0.0, 0.03, 0.06, 0.09, 0.12, 0.15, 0.18])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Service fee \n Unit flight cost = ' + str(cost_flightList[1]))
        plt.ylabel('Rate')
        
        
        ##
        ax = plt.subplot(233)
        avgReturnRate = []
        avgChargedRate = []
        avgProceedRate = []
        for k in range(len(cost_chargeFeeList)):
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            cnt = 0
            for i in range(len(self.data_Graph_Export_DP_ServiceFee_FlightCost[0])):    
                if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][1] == cost_flightList[2]):    
                    if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][2] == cost_chargeFeeList[k]):
                        sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][3]
                        sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][4]
                        cnt = cnt + 1
            
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            avgProceedRate.append(1 - (avgReturnRate[-1] + avgChargedRate[-1]))
        
        
        
        plt.plot(cost_chargeFeeList, avgReturnRate, '.--', c='blue', label='depot return rate')
        plt.plot(cost_chargeFeeList, avgChargedRate, '.--', c='black', label='getting charged via an EV rate')
        
        
        
        ax.set_xticks(cost_chargeFeeList)
        ax.set_yticks([0.0, 0.03, 0.06, 0.09, 0.12, 0.15, 0.18])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Service fee \n Unit flight cost = ' + str(cost_flightList[2]))
        plt.ylabel('Rate')
        
        
        ##
        ax = plt.subplot(234)
        avgReturnRate = []
        avgChargedRate = []
        avgProceedRate = []
        for k in range(len(cost_chargeFeeList)):
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            cnt = 0
            for i in range(len(self.data_Graph_Export_DP_ServiceFee_FlightCost[0])):    
                if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][1] == cost_flightList[3]):    
                    if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][2] == cost_chargeFeeList[k]):
                        sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][3]
                        sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][4]
                        cnt = cnt + 1
            
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            avgProceedRate.append(1 - (avgReturnRate[-1] + avgChargedRate[-1]))
        
        
        
        plt.plot(cost_chargeFeeList, avgReturnRate, '.--', c='blue', label='depot return rate')
        plt.plot(cost_chargeFeeList, avgChargedRate, '.--', c='black', label='getting charged via an EV rate')
        
        ax.set_xticks(cost_chargeFeeList)
        ax.set_yticks([0.0, 0.03, 0.06, 0.09, 0.12, 0.15, 0.18])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Service fee \n Unit flight cost = ' + str(cost_flightList[3]))
        plt.ylabel('Rate')
        
        
        ##
        ax = plt.subplot(235)
        avgReturnRate = []
        avgChargedRate = []
        avgProceedRate = []
        for k in range(len(cost_chargeFeeList)):
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            cnt = 0
            for i in range(len(self.data_Graph_Export_DP_ServiceFee_FlightCost[0])):    
                if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][1] == cost_flightList[4]):    
                    if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][2] == cost_chargeFeeList[k]):
                        sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][3]
                        sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][4]
                        cnt = cnt + 1
            
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            avgProceedRate.append(1 - (avgReturnRate[-1] + avgChargedRate[-1]))
        
        
        
        plt.plot(cost_chargeFeeList, avgReturnRate, '.--', c='blue', label='depot return rate')
        plt.plot(cost_chargeFeeList, avgChargedRate, '.--', c='black', label='getting charged via an EV rate')
        
        
        ax.set_xticks(cost_chargeFeeList)
        ax.set_yticks([0.0, 0.03, 0.06, 0.09, 0.12, 0.15, 0.18])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Service fee \n Unit flight cost = ' + str(cost_flightList[4]))
        plt.ylabel('Rate')
        
        
        ##
        ax = plt.subplot(236)
        avgReturnRate = []
        avgChargedRate = []
        avgProceedRate = []
        for k in range(len(cost_chargeFeeList)):
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            cnt = 0
            for i in range(len(self.data_Graph_Export_DP_ServiceFee_FlightCost[0])):    
                if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][1] == cost_flightList[5]):    
                    if(list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][2] == cost_chargeFeeList[k]):
                        sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][3]
                        sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_ServiceFee_FlightCost[:]))[i][0][4]
                        cnt = cnt + 1
            
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            avgProceedRate.append(1 - (avgReturnRate[-1] + avgChargedRate[-1]))
        
        
        
        plt.plot(cost_chargeFeeList, avgReturnRate, '.--', c='blue', label='depot return rate')
        plt.plot(cost_chargeFeeList, avgChargedRate, '.--', c='black', label='getting charged via an EV rate')
        
        
        ax.set_xticks(cost_chargeFeeList)
        ax.set_yticks([0.0, 0.03, 0.06, 0.09, 0.12, 0.15, 0.18])
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('Service fee \n Unit flight cost = ' + str(cost_flightList[5]))
        plt.ylabel('Rate')
        
        
        
        
        
        
        plt.legend(loc='best', bbox_to_anchor=(1, 1))
        plt.tight_layout()
        plt.show()   


    
    def Export_DP_UAVcapacity(self, UAVcapacityList, proceed_rate_avg, depot_return_rate_avg, charged_rate_avg, obj_DP_avg, obj_MIP_avg, proceed_rate_avg_MIP, depot_return_rate_avg_MIP, charged_rate_avg_MIP):
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        xlsFileName = "TempResult\Summary\DP_UAVCapacity" + ".xlsx"   

        ws["A1"] = "UAV Capacity"
        ws["B1"] = "[DP] Proceed rate(avg.)"
        ws["C1"] = "[DP] Depot return rate (avg.)"
        ws["D1"] = "[DP] Get charged rate (avg.)"
        ws["E1"] = "[DP] Obj. value (avg.)"
        ws["F1"] = "[MIP] Proceed rate(avg.)"
        ws["G1"] = "[MIP] Depot return rate (avg.)"
        ws["H1"] = "[MIP] Get charged rate (avg.)"
        ws["I1"] = "[MIP] Obj. value (avg.)"
        
        
        for l in range(len(UAVcapacityList)):
            ws.cell(row=(l + 2), column=1, value=UAVcapacityList[l])
            ws.cell(row=(l + 2), column=2, value=proceed_rate_avg[l])
            ws.cell(row=(l + 2), column=3, value=depot_return_rate_avg[l])
            ws.cell(row=(l + 2), column=4, value=charged_rate_avg[l])
            ws.cell(row=(l + 2), column=5, value=obj_DP_avg[l])
            
            ws.cell(row=(l + 2), column=6, value=proceed_rate_avg_MIP[l])
            ws.cell(row=(l + 2), column=7, value=depot_return_rate_avg_MIP[l])
            ws.cell(row=(l + 2), column=8, value=charged_rate_avg_MIP[l])
            ws.cell(row=(l + 2), column=9, value=obj_MIP_avg[l])
            
            tempArray = []
            tempArray.append(UAVcapacityList[l])
            tempArray.append(proceed_rate_avg[l])
            tempArray.append(depot_return_rate_avg[l])
            tempArray.append(charged_rate_avg[l])
            tempArray.append(obj_DP_avg[l])
            tempArray.append(proceed_rate_avg_MIP[l])
            tempArray.append(depot_return_rate_avg_MIP[l])
            tempArray.append(charged_rate_avg_MIP[l])
            tempArray.append(obj_MIP_avg[l])
            
            self.data_Graph_Export_DP_UAVCapacity.append(tempArray)
            
         
            
            
        wb.save(xlsFileName)
    
    
    

    
    def Export_DP_Arm_numEVs(self, arm_Tri_List, num_EV_List, proceed_rate_avg, depot_return_rate_avg, charged_rate_avg, obj_DP_avg):
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        xlsFileName = "TempResult\SensitivityAnalysis_Arm\Result" + ".xlsx"   

        ws["A1"] = "Arm"
        ws["B1"] = "Num EVs"
        ws["C1"] = "Proceed rate(avg.)"
        ws["D1"] = "Depot return rate (avg.)"
        ws["E1"] = "Get charged rate (avg.)"
        ws["F1"] = "Obj. value (avg.)"
        
        
        
        rowIndex = 0
        for k in range(len(arm_Tri_List)):
            
            
            for l in range(len(num_EV_List)):
                ws.cell(row=(rowIndex + 2), column=1, value=arm_Tri_List[k])
                ws.cell(row=(rowIndex + 2), column=2, value=num_EV_List[l])
                ws.cell(row=(rowIndex + 2), column=3, value=proceed_rate_avg[rowIndex])
                ws.cell(row=(rowIndex + 2), column=4, value=depot_return_rate_avg[rowIndex])
                ws.cell(row=(rowIndex + 2), column=5, value=charged_rate_avg[rowIndex])
                ws.cell(row=(rowIndex + 2), column=6, value=obj_DP_avg[rowIndex])
                
                tempArray = []
                tempArray.append(arm_Tri_List[k])
                tempArray.append(num_EV_List[l])
                tempArray.append(proceed_rate_avg[rowIndex])
                tempArray.append(depot_return_rate_avg[rowIndex])
                tempArray.append(charged_rate_avg[rowIndex])
                tempArray.append(obj_DP_avg[rowIndex])
                
                self.data_Graph_Export_DP_Arm_NumEVs.append(tempArray)
                
                rowIndex = rowIndex + 1
            
            
        wb.save(xlsFileName)

    
    
    def Graph_DP_UAVcapacity(self, UAVcapacityList, proceed_rate_avg, depot_return_rate_avg, charged_rate_avg, obj_DP_avg, obj_MIP_avg, proceed_rate_avg_MIP, depot_return_rate_avg_MIP, charged_rate_avg_MIP):
        
        avgProceedRate = []
        avgReturnRate = []
        avgChargedRate = []
        avgObjValue = []
        
        avgProceedRate_MIP = []
        avgReturnRate_MIP = []
        avgChargedRate_MIP = []
        avgObjValue_MIP = []
        
        
        for i in range(len(UAVcapacityList)):
            sumProceedRate = 0.0
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            sumObjValue = 0.0
            
            sumProceedRate_MIP = 0.0
            sumReturnRate_MIP = 0.0
            sumChargedRate_MIP = 0.0
            sumObjValue_MIP = 0.0
            
            cnt = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[0])):
                if(UAVcapacityList[i] == list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[0][j]):
                    sumProceedRate = sumProceedRate + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[1][j]
                    sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[2][j]
                    sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[3][j]
                    sumObjValue = sumObjValue + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[4][j]
                    
                    sumProceedRate_MIP = sumProceedRate_MIP + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[5][j]
                    sumReturnRate_MIP = sumReturnRate_MIP + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[6][j]
                    sumChargedRate_MIP = sumChargedRate_MIP + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[7][j]
                    sumObjValue_MIP = sumObjValue_MIP + list(zip(*self.data_Graph_Export_DP_UAVCapacity[:]))[8][j]
                    
                    cnt = cnt + 1
            
            avgProceedRate.append(sumProceedRate / cnt)
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            avgObjValue.append(sumObjValue / cnt)
            
            
            avgProceedRate_MIP.append(sumProceedRate_MIP / cnt)
            avgReturnRate_MIP.append(sumReturnRate_MIP / cnt)
            avgChargedRate_MIP.append(sumChargedRate_MIP / cnt)
            avgObjValue_MIP.append(sumObjValue_MIP / cnt)
            
        
        
        plt.rcParams['font.family'] = 'Times New Roman'
        plt.rc('font', size=10)
        
        
        plt.figure(1)
        
        
        ax = plt.subplot(121)    
        plt.plot(UAVcapacityList, avgProceedRate, '.', c='black', label='Avg. rate of proceeding to the next EV (Optimal policy)')
        plt.plot(UAVcapacityList, avgProceedRate_MIP, '.', c='gold', label='Avg. rate of proceeding to the next EV (Deterministic)')
        
        ax.set_xticks(UAVcapacityList)
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('UAV Capacity \n (a)')
        plt.ylabel('Rate')
        # plt.legend(loc='best', bbox_to_anchor=(1, 1))
        plt.legend(loc='best')
        
        
        
        ax = plt.subplot(122)
        
        plt.plot(UAVcapacityList, avgObjValue, '.', c='black', label='Avg. objective value (Optimal policy)')
        plt.plot(UAVcapacityList, avgObjValue_MIP, '.', c='gold', label='Avg. objective value (Deterministic)')
        
        ax.set_xticks(UAVcapacityList)
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        plt.xlabel('UAV Capacity \n (b)')
        plt.ylabel('Cost ($)')
        
        # plt.legend(loc='best', bbox_to_anchor=(1, 1))
        plt.legend(loc='best')
        plt.show()     
    
    
    
    
   
    def Graph_DP_Arm_numEVs(self, arm_Tri_List, num_EV_List, proceed_rate_avg, depot_return_rate_avg, charged_rate_avg, obj_DP_avg):
            
        plt.rcParams['font.family'] = 'Times New Roman'
        plt.rc('font', size=10)
        
        
        plt.figure(1)
        '''
        ax = plt.subplot(121)
        
        
        avgProceedRate = []
        avgReturnRate = []
        avgChargedRate = []
        
        for i in range(len(arm_Tri_List)):
            sumProceedRate = 0.0
            sumReturnRate = 0.0
            sumChargedRate = 0.0
            cnt = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0])):
                if(arm_Tri_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate = sumProceedRate + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    sumReturnRate = sumReturnRate + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[3][j]
                    sumChargedRate = sumChargedRate + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[4][j]
                    cnt = cnt + 1
            
            avgProceedRate.append(sumProceedRate / cnt)
            avgReturnRate.append(sumReturnRate / cnt)
            avgChargedRate.append(sumChargedRate / cnt)
            
        plt.plot(arm_Tri_List, avgProceedRate, '.', c='black', label='Avg. rate')
        
        ax.set_xticks(arm_Tri_List)
        ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        # plt.legend(loc='best')
        plt.xlabel('Arm \n (a)')
        plt.ylabel('Avg. Rate of proceeding to the next EV')
        
        '''
        
        
        
        # ax = plt.subplot(122)
        
        
        avgProceedRate_0 = []
        avgProceedRate_2 = []
        avgProceedRate_4 = []
        avgProceedRate_6 = []
        avgProceedRate_8 = []
        avgProceedRate_10 = []
        
        
        
        
        for i in range(len(num_EV_List)):
            sumProceedRate_0 = 0.0
            cnt_0 = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1])):
                if(num_EV_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1][j] and 0 == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate_0 = sumProceedRate_0 + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    cnt_0 = cnt_0 + 1
            avgProceedRate_0.append(sumProceedRate_0 / cnt_0)
            
            '''
            sumProceedRate_3 = 0.0
            cnt_3 = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1])):
                if(num_EV_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1][j] and 3 == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate_3 = sumProceedRate_3 + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    cnt_3 = cnt_3 + 1
            avgProceedRate_3.append(sumProceedRate_3 / cnt_3)
            '''
    
            
            
            sumProceedRate_2 = 0.0
            cnt_2 = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1])):
                if(num_EV_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1][j] and 4 == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate_2 = sumProceedRate_2 + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    cnt_2 = cnt_2 + 1
            avgProceedRate_2.append(sumProceedRate_2 / cnt_2)
            
            
            sumProceedRate_4 = 0.0
            cnt_4 = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1])):
                if(num_EV_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1][j] and 8 == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate_4 = sumProceedRate_4 + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    cnt_4 = cnt_4 + 1
            avgProceedRate_4.append(sumProceedRate_4 / cnt_4)
            
            
            sumProceedRate_6 = 0.0
            cnt_6 = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1])):
                if(num_EV_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1][j] and 12 == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate_6 = sumProceedRate_6 + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    cnt_6 = cnt_6 + 1
            avgProceedRate_6.append(sumProceedRate_6 / cnt_6)
            
            
            sumProceedRate_8 = 0.0
            cnt_8 = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1])):
                if(num_EV_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1][j] and 16 == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate_8 = sumProceedRate_8 + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    cnt_8 = cnt_8 + 1
            avgProceedRate_8.append(sumProceedRate_8 / cnt_8)
            
            
            sumProceedRate_10 = 0.0
            cnt_10 = 0
            for j in range(len(list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1])):
                if(num_EV_List[i] == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[1][j] and 20 == list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[0][j]):
                    sumProceedRate_10 = sumProceedRate_10 + list(zip(*self.data_Graph_Export_DP_Arm_NumEVs[:]))[2][j]
                    cnt_10 = cnt_10 + 1
            avgProceedRate_10.append(sumProceedRate_10 / cnt_10)
            

        color_map = plt.get_cmap("inferno")
        plt.plot(num_EV_List, avgProceedRate_0, '.--', c='blue', label='arm = 0')
        # plt.plot(num_EV_List, avgProceedRate_3, '.--', c='green', label='arm = 3')
        plt.plot(num_EV_List, avgProceedRate_2, '.--', c='green', label='arm = 4')
        # plt.plot(num_EV_List, avgProceedRate_9, '.--', c='cyan', label='arm = 9')
        plt.plot(num_EV_List, avgProceedRate_4, '.--', c='brown', label='arm = 8')
        # plt.plot(num_EV_List, avgProceedRate_15, '.--', c='yellow', label='arm = 15')
        plt.plot(num_EV_List, avgProceedRate_6, '.--', c='black', label='arm = 12')
        # plt.plot(num_EV_List, avgProceedRate_21, '.--', c='navy', label='arm = 21')
        plt.plot(num_EV_List, avgProceedRate_8, '.--', c='purple', label='arm = 16')
        plt.plot(num_EV_List, avgProceedRate_10, '.--', c='navy', label='arm = 20')
        
        '''
        plt.plot(num_EV_List, avgProceedRate_0, '.', c='red', label='Avg. rate of proceeding to the next EV, if arm = 0')
        plt.plot(num_EV_List, avgProceedRate_3, '.', c='orange', label='Avg. rate of proceeding to the next EV, if arm = 3')
        plt.plot(num_EV_List, avgProceedRate_6, '.', c='yellow', label='Avg. rate of proceeding to the next EV, if arm = 6')
        plt.plot(num_EV_List, avgProceedRate_9, '.', c='green', label='Avg. rate of proceeding to the next EV, if arm = 9')
        plt.plot(num_EV_List, avgProceedRate_12, '.', c='blue', label='Avg. rate of proceeding to the next EV, if arm = 12')
        plt.plot(num_EV_List, avgProceedRate_15, '.', c='navy', label='Avg. rate of proceeding to the next EV, if arm = 15')
        plt.plot(num_EV_List, avgProceedRate_18, '.', c='purple', label='Avg. rate of proceeding to the next EV, if arm = 18')
        plt.plot(num_EV_List, avgProceedRate_21, '.', c='black', label='Avg. rate of proceeding to the next EV, if arm = 21')
        plt.plot(num_EV_List, avgProceedRate_24, '.', c='gold', label='Avg. rate of proceeding to the next EV, if arm = 24')
        plt.plot(num_EV_List, avgProceedRate_27, '.', c='palegreen', label='Avg. rate of proceeding to the next EV, if arm = 27')
        '''
        
        
        
        # ax.set_xticks(num_EV_List)
        # ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        
        
        plt.xticks([2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
        plt.xlabel('Number of EVs')
        plt.ylabel('Avg. Rate of proceeding to the next EV')
        
        # plt.legend(loc='best', bbox_to_anchor=(1, 1))
        plt.legend(loc='best')
        plt.show()      
    
   
   
   
   
    
    def Export_DP_ServiceFee_FlightCost(self, armList, cost_chargeFeeList, cost_flightList, depot_return_rate_avg, charged_rate_avg, obj_DP_avg):
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        xlsFileName = "TempResult\SensitivityAnalysis_ServiceFee_FlightCost\Result" + ".xlsx"   

        ws["A1"] = "Arm"
        ws["B1"] = "Unit flight cost"
        ws["C1"] = "Service fee"
        ws["D1"] = "Depot return rate (avg.)"
        ws["E1"] = "Get charged rate (avg.)"
        ws["F1"] = "Obj. value (avg.)"
        
        
        rowIndex = 0
        for k in range(len(armList)):
            self.data_Graph_Export_DP_ServiceFee_FlightCost.append([])
            
            for l in range(len(cost_flightList)):
                for i in range(len(cost_chargeFeeList)):
                    ws.cell(row=(rowIndex + 2), column=1, value=armList[k])
                    ws.cell(row=(rowIndex + 2), column=2, value=cost_flightList[l])
                    ws.cell(row=(rowIndex + 2), column=3, value=cost_chargeFeeList[i])
                    ws.cell(row=(rowIndex + 2), column=4, value=depot_return_rate_avg[rowIndex])
                    ws.cell(row=(rowIndex + 2), column=5, value=charged_rate_avg[rowIndex])
                    ws.cell(row=(rowIndex + 2), column=6, value=obj_DP_avg[rowIndex])
                    
                    tempArray = []
                    tempArray.append(armList[k])
                    tempArray.append(cost_flightList[l])
                    tempArray.append(cost_chargeFeeList[i])
                    tempArray.append(depot_return_rate_avg[rowIndex])
                    tempArray.append(charged_rate_avg[rowIndex])
                    tempArray.append(obj_DP_avg[rowIndex])
                    
                    self.data_Graph_Export_DP_ServiceFee_FlightCost[k].append(tempArray)
                    
                    rowIndex = rowIndex + 1
            
        wb.save(xlsFileName)
    
    
    
        
    
    
    
        
    def Export_DP_Det_Comparison(self, arm_Tri_List, numEVList, obj_DP_avg, obj_MIP_avg, infeasible_avg, obj_Gap_avg, cpuTime_DP_avg, cpuTime_MIP_avg, numWorstCase_MIP_avg): 
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        xlsFileName = "TempResult\DP_Det_Comparison\Result" + ".xlsx"   

        ws["A2"] = "Arm"
        ws["B2"] = "Num of EVs"
        ws["C2"] = "DP policy obj. (avg.)"
        ws["D2"] = "DP policy CPU time (sec.) (avg.)"
        
        
        ws["F1"] = "Deterministic method"
        ws["F2"] = "obj. (avg.)"
        ws["G2"] = "Infeasible case (%)"
        ws["H2"] = "CPU time (sec.) (avg.)"
        
        ws["J2"] = "Num worst cases (avg.)"
        ws["K1"] = "Improvement in obj. of DP versus a Det method"
        
        
        
        for k in range(len(arm_Tri_List)):
            for i in range(len(numEVList)):
                ws.cell(row=(i+(k*len(numEVList))+3), column=1, value=arm_Tri_List[k])
                ws.cell(row=(i+(k*len(numEVList))+3), column=2, value=numEVList[i])
                ws.cell(row=(i+(k*len(numEVList))+3), column=3, value=obj_DP_avg[i+(k*len(numEVList))])
                ws.cell(row=(i+(k*len(numEVList))+3), column=4, value=cpuTime_DP_avg[i+(k*len(numEVList))])
                
                
                
                ws.cell(row=(i+(k*len(numEVList))+3), column=6, value=obj_MIP_avg[i+(k*len(numEVList))])
                ws.cell(row=(i+(k*len(numEVList))+3), column=7, value=infeasible_avg[i+(k*len(numEVList))])
                ws.cell(row=(i+(k*len(numEVList))+3), column=8, value=cpuTime_MIP_avg[i+(k*len(numEVList))])
                
                ws.cell(row=(i+(k*len(numEVList))+3), column=10, value=numWorstCase_MIP_avg[i+(k*len(numEVList))])
                
                ws.cell(row=(i+(k*len(numEVList))+3), column=11, value=obj_Gap_avg[i+(k*len(numEVList))])
            
            
        wb.save(xlsFileName)
                  
    
    
    def Export_DP_Policy_Comparison(self, arm_Tri_List, numEVList, obj_DP_avg, obj_DP_policy_avg, obj_Gap_avg, cpuTime_DP_avg, cpuTime_DP_policy_avg): 
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        xlsFileName = "TempResult\Summary\DP_Policy_Comparison" + ".xlsx"   

        ws["A2"] = "Arm"
        ws["B2"] = "Num of EVs"
        ws["C2"] = "DP obj. (avg.)"
        ws["D2"] = "DP CPU time (sec.) (avg.)"
        
        ws["E2"] = "DP policy obj. (avg.)"
        ws["F2"] = "DP policy CPU time (sec.) (avg.)"
        ws["G1"] = "Improvement in obj. of DP versus a DP policy"
        
        
        for k in range(len(arm_Tri_List)):
            for i in range(len(numEVList)):
                ws.cell(row=(i+(k*len(numEVList))+3), column=1, value=arm_Tri_List[k])
                ws.cell(row=(i+(k*len(numEVList))+3), column=2, value=numEVList[i])
                ws.cell(row=(i+(k*len(numEVList))+3), column=3, value=obj_DP_avg[i+(k*len(numEVList))])
                ws.cell(row=(i+(k*len(numEVList))+3), column=4, value=cpuTime_DP_avg[i+(k*len(numEVList))])
                
                ws.cell(row=(i+(k*len(numEVList))+3), column=5, value=obj_DP_policy_avg[i+(k*len(numEVList))])
                ws.cell(row=(i+(k*len(numEVList))+3), column=6, value=cpuTime_DP_policy_avg[i+(k*len(numEVList))]) 
                ws.cell(row=(i+(k*len(numEVList))+3), column=7, value=obj_Gap_avg[i+(k*len(numEVList))])
            
            
        wb.save(xlsFileName)
  
        
    
    
        
    def ExportOutputData(self, uavList, MIPFlag = True, PolicyFlag = True):
        # for DP
        wb = Workbook()
        ws = wb.active
        ws.title = "OutputData"
        
        if(PolicyFlag == False):
            xlsFileName = "TempResult\DP_DET_Comparison\DP_result_numEVs_" + str(self.numEVs) + "_" + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second) + str(datetime.now().microsecond) + ".xlsx"   
        else:
            xlsFileName = "TempResult\DP_DET_Comparison\DPpolicy_result_numEVs_" + str(self.numEVs) + "_" + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second) + str(datetime.now().microsecond) + ".xlsx"   

        ws["A1"] = "Obj. value_DP"
        ws["A2"] = "Obj. value_Policy"
        ws["A3"] = "DP algorithm CPU Time (sec)"
        
        if(PolicyFlag == False):
            ws["B1"] = self.objValue_DP
        else:    
            ws["B2"] = self.objValue_Policy
        
        if(PolicyFlag == False):
            ws["B3"] = self.CPUTime_DP
        else:
            ws["B3"] = self.CPUTime_Policy
        
        ws["A6"] = "Shortest distance (km) (depot-rendezvous)"
        ws["A7"] = "Longest distance (km) (depot-rendezvous)"
        (longest, shortest) = self.FindDistance()
        ws["B6"] = shortest
        ws["B7"] = longest
        
        
        ws["A9"] = "Charging stations"
        ws["B9"] = "Latitude"
        ws["C9"] = "Longitude"
        
        
        ws["A19"] = "EV"
        ws["B19"] = "Latitude"
        ws["C19"] = "Longitude"
        ws["D19"] = "Solution (0=proceed, 1=return, 2=charge)"
        ws["E19"] = "Realized discretized demand"
        ws["F19"] = "Discretized remaining energy"
        ws["G19"] = "Discretized flight energy to next rendezvous node from current node(x=0 or 2) or depot(x=1)"
        
        for i in range(len(self.chargingStations)):
            ws.cell(row=(i+10), column=1, value=i)
            ws.cell(row=(i+10), column=2, value=self.chargingStations[i].x)
            ws.cell(row=(i+10), column=3, value=self.chargingStations[i].y)
        
        
        for i in range(self.numEVs):
            ws.cell(row=(i+20), column=1, value=i)
            ws.cell(row=(i+20), column=2, value=self.evList[i].rendezNode.x)
            ws.cell(row=(i+20), column=3, value=self.evList[i].rendezNode.y)
            if(PolicyFlag == False):
                ws.cell(row=(i+20), column=4, value=self.solution_DP[i])
            else:
                ws.cell(row=(i+20), column=4, value=self.solution_Policy[i])
                
            ws.cell(row=(i+20), column=5, value=self.evList[i].realizedDemand)
            ws.cell(row=(i+20), column=6, value=self.realized_q_DP[i])
            if(self.solution_Policy[i] == 0):
                if(i == self.numEVs - 1):
                    ws.cell(row=(i+20), column=7, value=self.energyMatrix.Energy_3[i])
                else:
                    ws.cell(row=(i+20), column=7, value=self.energyMatrix.Energy_4[i, i+1])
                    
            elif(self.solution_Policy[i] == 1):
                ws.cell(row=(i+20), column=7, value=self.energyMatrix.Energy_2[i+1])
            
            elif(self.solution_Policy[i] == 2):
                if(i == self.numEVs - 1):
                    ws.cell(row=(i+20), column=7, value=self.energyMatrix.Energy_3[i])
                else:
                    ws.cell(row=(i+20), column=7, value=self.energyMatrix.Energy_4[i, i+1])
            
            
        wb.save(xlsFileName)


        if (MIPFlag):
            # for MIP
            wb = Workbook()
            ws = wb.active
            ws.title = "OutputData"
            xlsFileName = "TempResult\MIP_result_numEVs_" + str(self.numEVs) + "_" + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second) + str(datetime.now().microsecond) + ".xlsx"   

            ws["A1"] = "Obj. value_MIP"
            ws["B1"] = self.objValue_MIP
            
            ws["A3"] = "MIP CPU Time (sec)"
            ws["B3"] = self.CPUTime_MIP
            
            
            ws["A6"] = "Shortest distance (km) (depot-rendezvous)"
            ws["A7"] = "Longest distance (km) (depot-rendezvous)"
            (longest, shortest) = self.FindDistance()
            ws["B6"] = shortest
            ws["B7"] = longest
            
            
            ws["A9"] = "Charging stations"
            ws["B9"] = "Latitude"
            ws["C9"] = "Longitude"
            
            
            ws["A19"] = "EV"
            ws["B19"] = "Latitude"
            ws["C19"] = "Longitude"
            ws["D19"] = "x1"
            ws["E19"] = "x2"
            ws["F19"] = "y"
            ws["G19"] = "z"
            ws["H19"] = "Ddiscretized expected demand"
            ws["I19"] = "Discretized realized demand"
            ws["J19"] = "Discretized remaining energy"
            ws["K19"] = "Discretized flight energy to next rendezvous node from current node(x1=1 or z=1) or depot(x2=1 or y=1)"
            
            
            for i in range(len(self.chargingStations)):
                ws.cell(row=(i+10), column=1, value=i)
                ws.cell(row=(i+10), column=2, value=self.chargingStations[i].x)
                ws.cell(row=(i+10), column=3, value=self.chargingStations[i].y)
            
            for i in range(self.numEVs):
                ws.cell(row=(i+20), column=1, value=i)
                ws.cell(row=(i+20), column=2, value=self.evList[i].rendezNode.x)
                ws.cell(row=(i+20), column=3, value=self.evList[i].rendezNode.y)
                ws.cell(row=(i+20), column=4, value=self.x1_MIP[i])
                ws.cell(row=(i+20), column=5, value=self.x2_MIP[i])
                ws.cell(row=(i+20), column=6, value=self.y_MIP[i])
                ws.cell(row=(i+20), column=7, value=self.z_MIP[i])
                ws.cell(row=(i+20), column=8, value=self.evList[i].discretizedElectricityDemandCenter)
                ws.cell(row=(i+20), column=9, value=self.evList[i].realizedDemand)
                ws.cell(row=(i+20), column=10, value=self.realized_q_MIP[i])
                if(self.x1_MIP[i] == 1):
                    if(i == self.numEVs - 1):
                        ws.cell(row=(i+20), column=11, value=self.energyMatrix.Energy_3[i])
                    else:
                        ws.cell(row=(i+20), column=11, value=self.energyMatrix.Energy_4[i, i+1])
                        
                elif(self.x2_MIP[i] == 1):
                    ws.cell(row=(i+20), column=11, value=self.energyMatrix.Energy_2[i+1])
                        
                elif(self.y_MIP[i] == 1):
                    ws.cell(row=(i+20), column=11, value=self.energyMatrix.Energy_2[i+1])
                
                elif(self.z_MIP[i] == 1):
                    if(i == self.numEVs - 1):
                        ws.cell(row=(i+20), column=11, value=self.energyMatrix.Energy_3[i])
                    else:
                        ws.cell(row=(i+20), column=11, value=self.energyMatrix.Energy_4[i, i+1])
                    
                    
                    
            wb.save(xlsFileName)
    
        
    
    def ExportInputData(self, uav, numEVs):
        wb = Workbook()
        ws = wb.active
        ws.title = "InputParameter"
        xlsFileName = "TempResult\DP_DET_Comparison\InputParameter_numEVs_" + str(self.numEVs) + "_" + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second) + str(datetime.now().microsecond) + ".xlsx"   

        ws["A1"] = "UAV Capacity (kW·h)"
        ws["A2"] = "Route failure cost (cost/failure)"
        ws["A3"] = "Charge fee (cost/charge)"
        ws["A4"] = "Flight cost (cost/kW·h)"
        ws["A5"] = "Energy consumption rate (kW·h/km)"
        ws["A6"] = "Number of customers"
        
        
        ws["A13"] = "seed_EVsLocation"
        ws["A14"] = "seed_chargingStationLocation"
        
        ws["A16"] = "seed_EVsDemand"
        ws["A17"] = "arm_Tri (kW·h)"
        
        ws["A19"] = "map_UpLeft_Latitude"
        ws["A20"] = "map_UpLeft_Longitude"
        ws["A21"] = "map_UpRight_Latitude"
        ws["A22"] = "map_UpRight_Longitude"
        ws["A23"] = "map_DownRight_Latitude"
        ws["A24"] = "map_DownRight_Longitude"
        ws["A25"] = "map_DownLeft_Latitude"
        ws["A26"] = "map_DownLeft_Longitude"
        ws["A27"] = "depotLoc_Latitude"
        ws["A28"] = "depotLoc_Longitude"
        ws["A29"] = "M"
        
        
        ws["B1"] = uav.capacity 
        ws["B2"] = uav.routeFailureCost 
        ws["B3"] = uav.chargeFee 
        ws["B4"] = uav.unitFlightCost 
        ws["B5"] = uav.energyConsumptionRate 
        ws["B6"] = numEVs
        
        
        
        ws["B13"] = self.seed_EVsLocation 
        
        ws["B16"] = self.seed_EVsDemand 
        ws["B17"] = self.arm_Tri 
        
        ws["B19"] = self.map_UpLeft[0] 
        ws["B20"] = self.map_UpLeft[1] 
        ws["B21"] = self.map_UpRight[0] 
        ws["B22"] = self.map_UpRight[1] 
        ws["B23"] = self.map_DownRight[0] 
        ws["B24"] = self.map_DownRight[1] 
        ws["B25"] = self.map_DownLeft[0]
        ws["B26"] = self.map_DownLeft[1]
        ws["B27"] = self.depotLoc[0] 
        ws["B28"] = self.depotLoc[1] 
        ws["B29"] = self.M 
        
        
        wb.save(xlsFileName)
        
    
    
    
        
        
    
    
    