import os
from Node.Node import Node
os.environ["KMP_DUPLICATE_LIB_OK"]="TRUE"
import random
from tqdm import tqdm
from State.State import State
from Simulator.Sim import Sim
from Agent.EV import EV
from Agent.UAV import UAV
from Node.Node import Node
from Matrix.ODMatrix import ODMatrix
from Matrix.CostMatrix import CostMatrix
from Matrix.EnergyMatrix import EnergyMatrix
from Simulator.DynamicProgramming import DynamicPrograming
from Agent.MCS import MCS
from Matrix.ODMatrix import ODMatrix_MCS
from Matrix.EnergyMatrix import EnergyTimeMatrix_MCS
from openpyxl import Workbook
import math
import numpy as np


if __name__ == '__main__':
    
    ######################################### UAV Setting
    # Set the UAV's capacity (kW·h)
    uavCapacity = 1.56   
    
    # Set the UAV's routeFailureCost (cost/failure)
    cost_failure = 50 # default = 50
    
    # Set the UAV's chargeFee (cost/charge)
    cost_chargeFee = 5 # default = 5
    
    
    # Set unit UAV's flight cost (cost/kW·h)
    # cost_flight = 0.097
    cost_flight = 0.16  # default
    # cost_flight = 0.265
    
    # Set the UAV's energyConsumptionRate (kW·h/km)
    unitEnergyConsumption = 0.025405 
    
    # Set UAV speed
    uavSpeed = 10  # 10 m/s
    
    
    
    ######################################### MCS Setting
    # Set the MCS's capacity (kW·h)
    mcsCapacity = 16.4   
    
    
    
    #########################################
    
    # Set the EV speed per road type (km/h)
    speed0 = 60     # Expressway
    speed1 = 40     # Arterial road
    speed2 = 30     # Secondary trunk road
    speed3 = 20     # Branch road
    
    
    # Set the charging stations
    destinationList = []
    destinationList.append(Node(2, 37.77885, -122.41366))  # 1
    destinationList.append(Node(2, 37.78520, -122.40531))  # 2
    destinationList.append(Node(2, 37.78206, -122.40393))  # 3
    destinationList.append(Node(2, 37.78147, -122.39980))  # 4
    destinationList.append(Node(2, 37.78891, -122.39698))  # 5
    destinationList.append(Node(2, 37.78332, -122.39155))  # 6
    destinationList.append(Node(2, 37.77893, -122.39510))  # 7
    
    
    
    # Set the map (UpLeft) 
    # map_UpLeft = (37.778601, -122.414838)
    # map_UpLeft = (37.791024, -122.414838)
    # map_UpLeft = (37.791218, -122.421677)
    map_UpLeft = (37.790865, -122.4281007)
    
    # Set the map (UpRight) 
    map_UpRight = (37.791024, -122.399221)
    
    
    # (37.783903, -122.390829)
    
    
    map_UpRight = (37.791024, -122.390829)
    
    
    # Set the map (DownLeft) 
    # map_DownLeft = (37.769969, -122.404018)
    # map_DownLeft = (37.769969, -122.421677)
    map_DownLeft = (37.769969, -122.4281007)
    
    # Set the map (DownRight) 
    # map_DownRight = (37.782448, -122.388053)
    # map_DownRight = (37.769969, -122.388053)
    # map_DownRight = (37.769969, -122.399221)
    map_DownRight = (37.769969, -122.390829)
    
    # electricity를 discretize하는 단위
    unitDiscrete = 0.017
    
    # Set the random seed for customer locations
    seed_EVsLocation = 24224
    
    # Set the random seed for customer demand
    seed_EVsDemand = 312222
    
    # Set arm of discrete triangular distribution
    arm_Tri = 22
    
    # Set the number of customers
    numEVs = 50
    
    # Set the number of UAVs
    numUAVs = 1
    
    
    # Set the number of MCSs (실험 환경상 동일하게함, 고객 배분을 임의로 하기 때문에 서로 다른 배분을 UAV와 MCS에 줄 수 없음)
    numMCSs = numUAVs
    
    
    # Set the depot
    depotLoc = (37.760008, -122.382484)
    
    # Set the large M
    M = 100000000.0
    
    # Set charging rate
    chargingRate = 62   # kW
    
    
    # Set simulation start time
    startTime = 0.0
    
    # speed0
    UnitDistPowerConsumption_mean0 = 0.247 + (1.52 / speed0) - (0.004 * speed0) + (2.992 * 10**(-5) * speed0)
    # speed1
    UnitDistPowerConsumption_mean1 = -0.179 + (0.004 * speed1) + (5.492 / speed1)
    # speed2
    UnitDistPowerConsumption_mean2 = 0.21 - (0.001 * speed2) + (1.531 / speed2)
    # speed3
    UnitDistPowerConsumption_mean3 = 0.208 - (0.002 * speed3) + (1.553 / speed3)
    
    
    
    def ConvertDecimalDegreesToRadians(deg):
        return (deg * math.pi / 180)
        

    def ConvertRadiansToDecimalDegrees(rad):
        return (rad * 180 / math.pi)
    
    
    def GetKMDistance(p1, p2):
        theta = p1[1] - p2[1]
        dist = math.sin(ConvertDecimalDegreesToRadians(p1[0])) * math.sin(ConvertDecimalDegreesToRadians(p2[0])) + math.cos(ConvertDecimalDegreesToRadians(p1[0])) * math.cos(ConvertDecimalDegreesToRadians(p2[0])) * math.cos(ConvertDecimalDegreesToRadians(theta))
        dist = math.acos(dist)
        dist = ConvertRadiansToDecimalDegrees(dist)
        dist = dist * 60 * 1.1515
        dist = dist * 1.609344

        return dist
    
    def GetNearestChargingStation(rendezvousLoc):
        minDist = 999999999
        minIndex = 0
        for index, chargingStation in enumerate(destinationList):
            dist = GetKMDistance(rendezvousLoc, (chargingStation.x, chargingStation.y))
            if dist < minDist:
                minIndex = index
                minDist = dist
        return minIndex, minDist
    
    
    def GenerateEVs(numEVs, seed_EVsLocation, seed_EVsDemand, evList, groundVehicleDistanceWeight, arm_Tri):
        rndEVLocationGenerator = random.Random(seed_EVsLocation) 
        rndEVDemandGenerator = random.Random(seed_EVsDemand) 
        
        for j in range(numEVs):            
            
            # Set rendezvous node
            rendezvousLoc_y = rndEVLocationGenerator.uniform(map_UpLeft[1], map_UpRight[1])
            rendezvousLoc_x = rndEVLocationGenerator.uniform(map_UpLeft[0], map_DownLeft[0])

            
            # Set charging station (= destination)
            chargingStationIndex, r_d_distance = GetNearestChargingStation((rendezvousLoc_x, rendezvousLoc_y))
            r_d_distance = round(r_d_distance, 3)   # unit: km
            r_d_distance *= groundVehicleDistanceWeight # 가중치 곱함
            
            destination = destinationList[chargingStationIndex]
            
            
            # road type은 랜덤하게 선택 후 랜덤하게 배분
            speeds = []
            unitPowerConsumptionMeans = []  # unit: kWh/km
            for i in range(4):
                speedType = rndEVDemandGenerator.randint(0, 3)
                if speedType == 0:
                    speeds.append(speed0)
                    unitPowerConsumptionMeans.append(UnitDistPowerConsumption_mean0)
                elif speedType == 1:
                    speeds.append(speed1)
                    unitPowerConsumptionMeans.append(UnitDistPowerConsumption_mean1)
                elif speedType == 2:
                    speeds.append(speed2)
                    unitPowerConsumptionMeans.append(UnitDistPowerConsumption_mean2)
                elif speedType == 3:
                    speeds.append(speed3)
                    unitPowerConsumptionMeans.append(UnitDistPowerConsumption_mean3)
            
            
            distances = []
            totalDistance = int(round(r_d_distance * 1000000, 0))
            for i in range(4):
                if i == 3:
                    distance = round(totalDistance / 1000000, 3)
                    distances.append(distance)
                else:
                    distance = rndEVDemandGenerator.randrange(totalDistance)
                    distance = round(distance, 0)
                    totalDistance = totalDistance - distance
                    
                    distance = round(distance / 1000000, 3)
                    distances.append(distance)
                    
                
            # Discretized power consumption (= electricity demand) center
            totalPowerConsumption = 0.0
            for i in range(len(distances)):
                totalPowerConsumption += distances[i] * unitPowerConsumptionMeans[i]    # unit: kWh
                
                
            discretizedElectricityDemandCenter = State().Discretized(totalPowerConsumption)
            discretizedElectricityDemandMin = max(discretizedElectricityDemandCenter - arm_Tri, 1)
            discretizedElectricityDemandMax = discretizedElectricityDemandCenter + arm_Tri
            
            # Generate EV
            ev = EV(j, Node(1, rendezvousLoc_x, rendezvousLoc_y), destination, discretizedElectricityDemandCenter, discretizedElectricityDemandMin, discretizedElectricityDemandMax,
                    round(rndEVDemandGenerator.random() * 1000000), speeds, distances, arm_Tri)
            
            
            
            evList.append(ev)
    
    
    
    def InitializeUAVs(uavList):
        for uav in uavList:
            uav.remainingEnergyList = []
            uav.scheduleList = []
            uav.solution_DP = None
            uav.solution_policy = None
            uav.h = np.zeros(len(uav.evList))
            uav.objValue_DP = 0
            uav.ObjValue_policy = 0
            uav.remainingEnergy = uav.capacity
            # Define DP
            uav.dp = DynamicPrograming(uav, M)
            
    def InitializeMCSs(mcsList):
        for mcs in mcsList:
            mcs.remainingEnergyList = []
            mcs.scheduleList = []
            mcs.objValue = 0
            mcs.remainingEnergy = mcs.capacity
    
    
    
    def ExportSensitivityAnalysis_NumUAV_MCS_EV(result_list):
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        xlsFileName = "TempResult\SensitivityAnalysis_UAV_MCS\Result" + ".xlsx"   

        ws["A1"] = "num EVs"
        ws["B1"] = "num UAVs"
        ws["C1"] = "num MCSs"
        ws["D1"] = "avg. completion time (UAV)"
        ws["E1"] = "avg. completion time (MCS)"
        
        
        for l in range(len(result_list)):
            ws.cell(row=(l + 2), column=1, value=result_list[l]['numEVs'])
            ws.cell(row=(l + 2), column=2, value=result_list[l]['numUAVs'])
            ws.cell(row=(l + 2), column=3, value=result_list[l]['numMCSs'])
            ws.cell(row=(l + 2), column=4, value=result_list[l]['avgCompletionTime_UAV'])
            ws.cell(row=(l + 2), column=5, value=result_list[l]['avgCompletionTime_MCS'])
            
            
        wb.save(xlsFileName)
        
    
    #################################################################################################
    # Set the types of experiments
    experimentType = 7  # 0: default, 1: DP vs. Deterministic, 2: service fee vs. flight cost, 3: Arm and num of EVs,
                        # 7: charging demand variation (UAV vs. MCS)
    
    # Set the random seed for random experiments
    seed_experiments = 233434
    
    
    if experimentType == 0: # default
        
        groundVehicleDistanceWeight = 1.5
        
        # Set arm of discrete triangular distribution
        arm_Tri = 2
        # Set the number of customers
        numEVs = 5
        
        numUAVs = 1
        numMCSs = 1
        
        
        ######################## Problem setting
        # Generate State
        state = State()
        
        
        # EV를 UAV 별로 몇개씩 할당할지 결정 (UAV 당 최소 EV 1개씩은 할당시킴)
        rndEVDivideGenerator = random.Random(seed_EVsLocation) 
        
        numAssignedEVs_UAV = []
        cumNumEVs = 0
        emptyUAVs = 0
        for i in range(numUAVs):
            if cumNumEVs == numEVs:
                numAssignedEVs_UAV.append(0)
                emptyUAVs += 1
            else:
                numAssignedEVs_UAV.append(1)
                cumNumEVs += 1
        
        remainingNumEVs = numEVs - cumNumEVs
            
        for i in range(numUAVs):
            if numAssignedEVs_UAV[i] != 0:
                if i == numUAVs - emptyUAVs - 1:
                    numAssignedEVs_UAV[i] += remainingNumEVs
                    break
                numAssignedEVs = rndEVDivideGenerator.randint(min(1, remainingNumEVs), remainingNumEVs)
                remainingNumEVs -= numAssignedEVs
                numAssignedEVs_UAV[i] += numAssignedEVs
        
            
        # Generate and assign EVs (customers)
        evList = []
        GenerateEVs(numEVs, seed_EVsLocation, seed_EVsDemand, evList, groundVehicleDistanceWeight, arm_Tri)
        
        
        # EV를 UAV 별로 할당하기 위한 list 준비
        evList_UAVs = []
        lastIndex = 0
        cumNumAssignedEVs = 0 
        for numAssignedEVs in numAssignedEVs_UAV:
            tempEVList = []
            tempEVList.extend(evList[lastIndex : lastIndex + numAssignedEVs])
            evList_UAVs.append(tempEVList)
            cumNumAssignedEVs += numAssignedEVs
            lastIndex = cumNumAssignedEVs
    
    
        # EV의 index를 UAV별로 0부터 시작하게 만들기 (OD matrix & energy matrix와 연관있음)
        for evList_UAV in evList_UAVs:
            for index, ev in enumerate(evList_UAV):
                ev.index = index
    
        uavList = []
        mcsList = []
        for i in range(numUAVs - emptyUAVs):
            #### Generate an UAV
            uav = UAV(cost_failure, cost_chargeFee, cost_flight, unitEnergyConsumption, uavCapacity, state, numEVs, chargingRate, startTime, uavSpeed, evList_UAVs[i], unitDiscrete)
            # Set the depot
            uav.setDepot(depotLoc[0], depotLoc[1])
            
            
            # Generate OD matrices
            uav.odMatrix = ODMatrix(uav.depot, uav.evList)
            
            # Generate Energy consumption matrices
            uav.energyMatrix = EnergyMatrix(uav.energyConsumptionRate, uav.odMatrix, uav.evList, state)
            
            # Generate Cost matrices
            uav.costMatrix = CostMatrix(uav.unitFlightCost, uav.energyMatrix, uav.evList)
            
            # Define DP
            uav.dp = DynamicPrograming(uav, M)
            
            uavList.append(uav)

            
            #### Generate an MCS
            mcs = MCS(mcsCapacity, state, chargingRate, startTime, evList_UAVs[i], unitDiscrete)
            # Set the depot
            mcs.setDepot(depotLoc[0], depotLoc[1])
            
            # Generate OD matrices
            mcs.odMatrix = ODMatrix_MCS(mcs.depot, uav.evList, groundVehicleDistanceWeight)
            
            # Generate Energy consumption matrices
            mcs.energyTimeMatrix = EnergyTimeMatrix_MCS(mcs.odMatrix, mcs.evList, state, seed_EVsDemand, UnitDistPowerConsumption_mean0, UnitDistPowerConsumption_mean1,
                                                UnitDistPowerConsumption_mean2, UnitDistPowerConsumption_mean3, speed0, speed1, speed2, speed3, arm_Tri)
            
            
            mcsList.append(mcs)
        
        
        result_list = []
        
        # Sim
        sim = Sim(M, uavList, mcsList, numEVs, result_list, len(uavList) + emptyUAVs, False, True, False, True)    
        
        # Export Input data
        # sim.ExportInputData()    
        
        # Export Output data
        # sim.ExportOutputData(False, True)
        
        
    elif experimentType == 1: # DP vs. Deterministic
        
        # Set the number of experiments    
        trial = 100
        # trial = 1
        
        
        rndExperimentsGenerator = random.Random(seed_experiments) 
        # numEVList = [2, 3, 4, 5, 6, 7, 8, 9, 10]
        numEVList = [2, 3, 4, 5, 6, 7, 8, 9, 10]
        arm_Tri_List = [1, 2, 3, 4, 5]
        
        obj_DP_avg = []
        obj_MIP_avg = []
        infeasible_avg = []
        obj_Gap_avg = []
        cpuTime_DP_avg = []
        cpuTime_MIP_avg = []
        avgCompletionTime_DP_avg = []
        avgCompletionTime_MIP_avg = []
        numWorstCase_MIP_avg = []
        
        # Generate State
        state = State()
        
        numUAVs = 1
        
        groundVehicleDistanceWeight = 1.5
        
        
        
        
        
        for k in range(len(arm_Tri_List)):
            arm_Tri = arm_Tri_List[k]
            
            for i in range(len(numEVList)):
                numEVs = numEVList[i]
                
                obj_DP = []
                obj_MIP = []
                obj_Gap = []
                infeasible_MIP = []
                cpuTime_DP = []
                cpuTime_MIP = []
                avgCompletionTime_DP = []
                avgCompletionTime_MIP = []
                numWorstCase_MIP = []
                
                for j in range(trial):
                    
                    seed_EVsLocation = seed_EVsLocation + int(round(rndExperimentsGenerator.random() * 10000))
                    seed_EVsDemand = seed_EVsDemand + int(round(rndExperimentsGenerator.random() * 10000)) 
                    
                    
                    
                    ######################################################
                    ######################## Problem setting
                    
                    
                   
                    
                    
                    
                    # EV를 UAV 별로 몇개씩 할당할지 결정 (UAV 당 최소 EV 1개씩은 할당시킴)
                    rndEVDivideGenerator = random.Random(seed_EVsLocation) 
                    
                    numAssignedEVs_UAV = []
                    cumNumEVs = 0
                    emptyUAVs = 0
                    for i in range(numUAVs):
                        if cumNumEVs == numEVs:
                            numAssignedEVs_UAV.append(0)
                            emptyUAVs += 1
                        else:
                            numAssignedEVs_UAV.append(1)
                            cumNumEVs += 1
                    
                    remainingNumEVs = numEVs - cumNumEVs
                        
                    for i in range(numUAVs):
                        if numAssignedEVs_UAV[i] != 0:
                            if i == numUAVs - emptyUAVs - 1:
                                numAssignedEVs_UAV[i] += remainingNumEVs
                                break
                            numAssignedEVs = rndEVDivideGenerator.randint(min(1, remainingNumEVs), remainingNumEVs)
                            remainingNumEVs -= numAssignedEVs
                            numAssignedEVs_UAV[i] += numAssignedEVs
                    
                        
                    # Generate and assign EVs (customers)
                    evList = []
                    GenerateEVs(numEVs, seed_EVsLocation, seed_EVsDemand, evList, groundVehicleDistanceWeight, arm_Tri)
                    
                    
                    # EV를 UAV 별로 할당하기 위한 list 준비
                    evList_UAVs = []
                    lastIndex = 0
                    cumNumAssignedEVs = 0 
                    for numAssignedEVs in numAssignedEVs_UAV:
                        tempEVList = []
                        tempEVList.extend(evList[lastIndex : lastIndex + numAssignedEVs])
                        evList_UAVs.append(tempEVList)
                        cumNumAssignedEVs += numAssignedEVs
                        lastIndex = cumNumAssignedEVs
                
                
                    # EV의 index를 UAV별로 0부터 시작하게 만들기 (OD matrix & energy matrix와 연관있음)
                    for evList_UAV in evList_UAVs:
                        for index, ev in enumerate(evList_UAV):
                            ev.index = index
                
                    uavList = []
                    
                    for i in range(numUAVs - emptyUAVs):
                        #### Generate an UAV
                        uav = UAV(cost_failure, cost_chargeFee, cost_flight, unitEnergyConsumption, uavCapacity, state, numEVs, chargingRate, startTime, uavSpeed, evList_UAVs[i], unitDiscrete)
                        # Set the depot
                        uav.setDepot(depotLoc[0], depotLoc[1])
                        
                        
                        # Generate OD matrices
                        uav.odMatrix = ODMatrix(uav.depot, uav.evList)
                        
                        # Generate Energy consumption matrices
                        uav.energyMatrix = EnergyMatrix(uav.energyConsumptionRate, uav.odMatrix, uav.evList, state)
                        
                        # Generate Cost matrices
                        uav.costMatrix = CostMatrix(uav.unitFlightCost, uav.energyMatrix, uav.evList)
                        
                        # Define DP
                        uav.dp = DynamicPrograming(uav, M)
                        
                        uavList.append(uav)

                    
                    result_list = []
                    
                    
                   
                    
                    
                    # Sim
                    sim = Sim(M, uavList, None, numEVs, result_list, len(uavList) + emptyUAVs, True, True) 
                    
                    
                    if uav.feasible_MIP == True:
                        obj_DP.append(uav.objValue_Policy)
                        obj_MIP.append(uav.objValue_MIP)
                        obj_Gap.append(obj_MIP[len(obj_MIP) - 1] - obj_DP[len(obj_DP) - 1])
                        infeasible_MIP.append(uav.feasible_MIP)
                        cpuTime_DP.append(uav.CPUTime_Policy)
                        cpuTime_MIP.append(uav.CPUTime_MIP)
                        numWorstCase_MIP.append(result_list[-1]["numWorstCase"])
                        
                    else:
                        obj_DP.append(0)
                        obj_MIP.append(0)
                        obj_Gap.append(0)
                        infeasible_MIP.append(0)
                        cpuTime_DP.append(0)
                        cpuTime_MIP.append(0)
                        numWorstCase_MIP.append(0)
                    
                    # Export Input data
                    # sim.ExportInputData(uavList[0], numEVs)    
                    
                    # Export Output data
                    # sim.ExportOutputData(uavList, True, True)

                    '''
                    # test start
                    if round(uav.objValue_Policy, 2) > round(uav.objValue_MIP, 2) and numEVs == 2:
                        print('test')
                    # test end
                    
                    # test start
                    if round(uav.objValue_Policy, 2) < round(uav.objValue_MIP, 2):
                        print('test')
                    # test end
                    '''
            
                # Save results
                sum_obj_DP = 0.0
                sum_obj_MIP = 0.0
                infeasible_MIP_cnt = 0
                sum_obj_Gap = 0.0
                sum_cpuTime_Policy = 0.0
                sum_cpuTime_MIP = 0.0
                sum_numWorstCase_MIP = 0.0
                
                for j in range(len(obj_DP)):
                    
                    if (infeasible_MIP[j] == False):
                        infeasible_MIP_cnt = infeasible_MIP_cnt + 1
                    else:
                        sum_obj_DP = sum_obj_DP + obj_DP[j]
                        sum_obj_MIP = sum_obj_MIP + obj_MIP[j]
                        sum_obj_Gap = sum_obj_Gap + obj_Gap[j]
                        sum_cpuTime_Policy = sum_cpuTime_Policy + cpuTime_DP[j]
                        sum_cpuTime_MIP = sum_cpuTime_MIP + cpuTime_MIP[j]
                        sum_numWorstCase_MIP = sum_numWorstCase_MIP + numWorstCase_MIP[j]
                
                
                obj_DP_avg.append(round(sum_obj_DP / (len(obj_DP) - infeasible_MIP_cnt), 2))
                obj_MIP_avg.append(round(sum_obj_MIP / (len(obj_DP) - infeasible_MIP_cnt), 2))
                infeasible_avg.append(round((infeasible_MIP_cnt / len(obj_DP)) * 100.0, 2))   
                obj_Gap_avg.append(round(sum_obj_Gap / (len(obj_DP) - infeasible_MIP_cnt), 2))
                cpuTime_DP_avg.append(round(sum_cpuTime_Policy / (len(obj_DP) - infeasible_MIP_cnt), 2))
                cpuTime_MIP_avg.append(round(sum_cpuTime_MIP / (len(obj_DP) - infeasible_MIP_cnt), 2))
                
                numWorstCase_MIP_avg.append(round(sum_numWorstCase_MIP / (len(obj_DP) - infeasible_MIP_cnt), 2))
                    
                    
        # Export results
        sim.Export_DP_Det_Comparison(arm_Tri_List, numEVList, obj_DP_avg, obj_MIP_avg, infeasible_avg, obj_Gap_avg, cpuTime_DP_avg, cpuTime_MIP_avg, numWorstCase_MIP_avg)
      
    
    elif experimentType == 2: # service fee vs. flight cost
        rndExperimentsGenerator = random.Random(seed_experiments) 
        armList = [2] 
        cost_flightList = [0.08, 0.16, 0.19, 0.23, 0.33, 0.48]  # 0.08 = China, 0.16 = United States, 0.19 = France, 0.23 = Japan, 0.33 = United Kingdom, 0.48 = Denmark
        cost_chargeFeeList = [1, 2, 3, 4, 5]
        
        depot_return_rate_avg = []
        charged_rate_avg = []
        obj_DP_avg = []
        numEVs = 10
        groundVehicleDistanceWeight = 1.5
        
        trial = 100
        
        
        # Generate State
        state = State()
        
        for k in range(len(armList)):
            
            arm_Tri = armList[k]
            for l in range(len(cost_flightList)):
                cost_flight = cost_flightList[l]
                
                for i in range(len(cost_chargeFeeList)):
                    cost_chargeFee = cost_chargeFeeList[i]
                    
                    depot_return_rate = []
                    charged_rate = []
                    obj_DP = []
                                    
                    for j in range(trial):
                        
                        seed_EVsLocation = seed_EVsLocation + int(round(rndExperimentsGenerator.random() * 1000))
                        seed_EVsDemand = seed_EVsDemand + int(round(rndExperimentsGenerator.random() * 1000)) 
                        
                        
                        
                        ######################## Problem setting
                        # EV를 UAV 별로 몇개씩 할당할지 결정 (UAV 당 최소 EV 1개씩은 할당시킴)
                        rndEVDivideGenerator = random.Random(seed_EVsLocation) 
                        
                        numAssignedEVs_UAV = []
                        cumNumEVs = 0
                        emptyUAVs = 0
                        for i in range(numUAVs):
                            if cumNumEVs == numEVs:
                                numAssignedEVs_UAV.append(0)
                                emptyUAVs += 1
                            else:
                                numAssignedEVs_UAV.append(1)
                                cumNumEVs += 1
                        
                        remainingNumEVs = numEVs - cumNumEVs
                            
                        for i in range(numUAVs):
                            if numAssignedEVs_UAV[i] != 0:
                                if i == numUAVs - emptyUAVs - 1:
                                    numAssignedEVs_UAV[i] += remainingNumEVs
                                    break
                                numAssignedEVs = rndEVDivideGenerator.randint(min(1, remainingNumEVs), remainingNumEVs)
                                remainingNumEVs -= numAssignedEVs
                                numAssignedEVs_UAV[i] += numAssignedEVs
                        
                            
                        # Generate and assign EVs (customers)
                        evList = []
                        GenerateEVs(numEVs, seed_EVsLocation, seed_EVsDemand, evList, groundVehicleDistanceWeight, arm_Tri)
                        
                        
                        # EV를 UAV 별로 할당하기 위한 list 준비
                        evList_UAVs = []
                        lastIndex = 0
                        cumNumAssignedEVs = 0 
                        for numAssignedEVs in numAssignedEVs_UAV:
                            tempEVList = []
                            tempEVList.extend(evList[lastIndex : lastIndex + numAssignedEVs])
                            evList_UAVs.append(tempEVList)
                            cumNumAssignedEVs += numAssignedEVs
                            lastIndex = cumNumAssignedEVs
                    
                    
                        # EV의 index를 UAV별로 0부터 시작하게 만들기 (OD matrix & energy matrix와 연관있음)
                        for evList_UAV in evList_UAVs:
                            for index, ev in enumerate(evList_UAV):
                                ev.index = index
                    
                        uavList = []
                        
                        for i in range(numUAVs - emptyUAVs):
                            #### Generate an UAV
                            uav = UAV(cost_failure, cost_chargeFee, cost_flight, unitEnergyConsumption, uavCapacity, state, numEVs, chargingRate, startTime, uavSpeed, evList_UAVs[i], unitDiscrete)
                            # Set the depot
                            uav.setDepot(depotLoc[0], depotLoc[1])
                            
                            
                            # Generate OD matrices
                            uav.odMatrix = ODMatrix(uav.depot, uav.evList)
                            
                            # Generate Energy consumption matrices
                            uav.energyMatrix = EnergyMatrix(uav.energyConsumptionRate, uav.odMatrix, uav.evList, state)
                            
                            # Generate Cost matrices
                            uav.costMatrix = CostMatrix(uav.unitFlightCost, uav.energyMatrix, uav.evList)
                            
                            # Define DP
                            uav.dp = DynamicPrograming(uav, M)
                            
                            uavList.append(uav)

                        
                        result_list = []
                        # Sim
                        sim = Sim(M, uavList, None, numEVs, result_list, len(uavList) + emptyUAVs, False, True) 
                        #################################################
                        
                        
                        obj_DP.append(uav.objValue_Policy)
                        
                        
                        cnt_depot_return = 0
                        cnt_charged = 0
                        for sol in range(len(uav.solution_policy)):
                            if(uav.solution_policy[sol] == 1):
                                cnt_depot_return = cnt_depot_return + 1
                            elif(uav.solution_policy[sol] == 2):
                                cnt_charged = cnt_charged + 1
                            
                        
                        depot_return_rate.append(round(cnt_depot_return / len(uav.solution_policy), 2))
                        charged_rate.append(round(cnt_charged / len(uav.solution_policy), 2))
                        
                        
                        # Export Input data
                        # sim.ExportInputData()    
                        
                        # Export Output data
                        # sim.ExportOutputData(False, True)
            
                
                    # Save results
                    sum_depot_return_rate = 0.0
                    sum_charged_rate = 0.0
                    sum_obj_DP = 0.0
                    
                    for j in range(len(depot_return_rate)):
                        sum_obj_DP = sum_obj_DP + obj_DP[j]
                        sum_depot_return_rate = sum_depot_return_rate + depot_return_rate[j]
                        sum_charged_rate = sum_charged_rate + charged_rate[j]
                    
                    
                    depot_return_rate_avg.append(round(sum_depot_return_rate / len(depot_return_rate), 2))
                    charged_rate_avg.append(round(sum_charged_rate / len(depot_return_rate), 2))
                    obj_DP_avg.append(round(sum_obj_DP / len(obj_DP), 2))
                
                    
        # Export results
        sim.Export_DP_ServiceFee_FlightCost(armList, cost_chargeFeeList, cost_flightList, depot_return_rate_avg, charged_rate_avg, obj_DP_avg)            
        
        # Export graphs
        sim.Graph_DP_ServiceFee_FlightCost(armList, cost_chargeFeeList, cost_flightList, depot_return_rate_avg, charged_rate_avg, obj_DP_avg) 
    
    
    elif experimentType == 3: # Arm and num of EVs
        rndExperimentsGenerator = random.Random(seed_experiments) 
        arm_Tri_List = [0, 4, 8, 12, 16, 20]
        num_EV_List = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
        
        
        proceed_rate_avg = []
        depot_return_rate_avg = []
        charged_rate_avg = []
        obj_DP_avg = []
        
        
        # Generate State
        state = State()
        
        groundVehicleDistanceWeight = 1.5
        
        trial = 100
        
        
        
        for k in range(len(arm_Tri_List)):
            
            arm_Tri = arm_Tri_List[k]
            for l in range(len(num_EV_List)):
                numEVs = num_EV_List[l]
                
                proceed_rate = []    
                depot_return_rate = []
                charged_rate = []
                obj_DP = []
                                
                for j in range(trial):
                    
                    seed_EVsLocation = seed_EVsLocation + int(round(rndExperimentsGenerator.random() * 1000))
                    seed_EVsDemand = seed_EVsDemand + int(round(rndExperimentsGenerator.random() * 1000)) 
                    
                    
                    
                    ####################################################################
                    ######################## Problem setting
                    
                    
                        
                    # Generate and assign EVs (customers)
                    evList = []
                    GenerateEVs(numEVs, seed_EVsLocation, seed_EVsDemand, evList, groundVehicleDistanceWeight, arm_Tri)
                    
                    
                    
                    
                    
                    # EV를 UAV 별로 몇개씩 할당할지 결정 (UAV 당 최소 EV 1개씩은 할당시킴)
                    rndEVDivideGenerator = random.Random(seed_EVsLocation) 
                    
                    numAssignedEVs_UAV = []
                    cumNumEVs = 0
                    emptyUAVs = 0
                    for i in range(numUAVs):
                        if cumNumEVs == numEVs:
                            numAssignedEVs_UAV.append(0)
                            emptyUAVs += 1
                        else:
                            numAssignedEVs_UAV.append(1)
                            cumNumEVs += 1
                    
                    remainingNumEVs = numEVs - cumNumEVs
                        
                    for i in range(numUAVs):
                        if numAssignedEVs_UAV[i] != 0:
                            if i == numUAVs - emptyUAVs - 1:
                                numAssignedEVs_UAV[i] += remainingNumEVs
                                break
                            numAssignedEVs = rndEVDivideGenerator.randint(min(1, remainingNumEVs), remainingNumEVs)
                            remainingNumEVs -= numAssignedEVs
                            numAssignedEVs_UAV[i] += numAssignedEVs
                    
                
                
                
                    # EV를 UAV 별로 할당하기 위한 list 준비
                    evList_UAVs = []
                    lastIndex = 0
                    cumNumAssignedEVs = 0 
                    for numAssignedEVs in numAssignedEVs_UAV:
                        tempEVList = []
                        tempEVList.extend(evList[lastIndex : lastIndex + numAssignedEVs])
                        evList_UAVs.append(tempEVList)
                        cumNumAssignedEVs += numAssignedEVs
                        lastIndex = cumNumAssignedEVs
                
                
                    # EV의 index를 UAV별로 0부터 시작하게 만들기 (OD matrix & energy matrix와 연관있음)
                    for evList_UAV in evList_UAVs:
                        for index, ev in enumerate(evList_UAV):
                            ev.index = index
                
                    uavList = []
                    
                    for i in range(numUAVs - emptyUAVs):
                        #### Generate an UAV
                        uav = UAV(cost_failure, cost_chargeFee, cost_flight, unitEnergyConsumption, uavCapacity, state, numEVs, chargingRate, startTime, uavSpeed, evList_UAVs[i], unitDiscrete)
                        # Set the depot
                        uav.setDepot(depotLoc[0], depotLoc[1])
                        
                        
                        # Generate OD matrices
                        uav.odMatrix = ODMatrix(uav.depot, uav.evList)
                        
                        # Generate Energy consumption matrices
                        uav.energyMatrix = EnergyMatrix(uav.energyConsumptionRate, uav.odMatrix, uav.evList, state)
                        
                        # Generate Cost matrices
                        uav.costMatrix = CostMatrix(uav.unitFlightCost, uav.energyMatrix, uav.evList)
                        
                        # Define DP
                        uav.dp = DynamicPrograming(uav, M)
                        
                        uavList.append(uav)

                    
                    result_list = []
                    # Sim
                    sim = Sim(M, uavList, None, numEVs, result_list, len(uavList) + emptyUAVs, False, True) 
                    
                    ####################################################################
                    

                    
                    obj_DP.append(uav.objValue_Policy)
                    
                    cnt_proceed = 0
                    cnt_depot_return = 0
                    cnt_charged = 0
                    for sol in range(len(uav.solution_policy)):
                        if(uav.solution_policy[sol] == 1):
                            cnt_depot_return = cnt_depot_return + 1
                        elif(uav.solution_policy[sol] == 2):
                            cnt_charged = cnt_charged + 1
                        elif(uav.solution_policy[sol] == 0):
                            cnt_proceed = cnt_proceed + 1
                        
                    proceed_rate.append(round(cnt_proceed / len(uav.solution_policy), 2))
                    depot_return_rate.append(round(cnt_depot_return / len(uav.solution_policy), 2))
                    charged_rate.append(round(cnt_charged / len(uav.solution_policy), 2))
                    
                    
                    # Export Input data
                    # sim.ExportInputData()    
                    
                    # Export Output data
                    # sim.ExportOutputData(False, True)
        
            
                # Save results
                sum_proceed_rate = 0.0
                sum_depot_return_rate = 0.0
                sum_charged_rate = 0.0
                sum_obj_DP = 0.0
                
                for j in range(len(depot_return_rate)):
                    sum_obj_DP = sum_obj_DP + obj_DP[j]
                    sum_depot_return_rate = sum_depot_return_rate + depot_return_rate[j]
                    sum_charged_rate = sum_charged_rate + charged_rate[j]
                    sum_proceed_rate = sum_proceed_rate + proceed_rate[j]
                
                
                proceed_rate_avg.append(round(sum_proceed_rate / len(depot_return_rate), 2))
                depot_return_rate_avg.append(round(sum_depot_return_rate / len(depot_return_rate), 2))
                charged_rate_avg.append(round(sum_charged_rate / len(depot_return_rate), 2))
                obj_DP_avg.append(round(sum_obj_DP / len(obj_DP), 2))
                
                    
        # Export results
        sim.Export_DP_Arm_numEVs(arm_Tri_List, num_EV_List, proceed_rate_avg, depot_return_rate_avg, charged_rate_avg, obj_DP_avg)            
        
        # Export graphs
        sim.Graph_DP_Arm_numEVs(arm_Tri_List, num_EV_List, proceed_rate_avg, depot_return_rate_avg, charged_rate_avg, obj_DP_avg) 
        
        
    
    
    elif experimentType == 6: # UAV vs. MCS    
        
        # Set arm of discrete triangular distribution
        arm_Tri = 2
        # Set the number of customers
        numEVs = 5
        
        # Sim
        sim = Sim(cost_failure, cost_chargeFee, cost_flight, unitEnergyConsumption, uavCapacity, numEVs, destinationList, 
              seed_EVsLocation, seed_EVsDemand, arm_Tri,
              map_UpLeft, map_UpRight, map_DownRight, map_DownLeft, depotLoc, speed0, speed1, speed2, speed3, M, chargingRate, startTime, uavSpeed, 
              numUAVs, numMCSs, mcsCapacity, unitDiscrete, distanceWeight, False, True, False, True)    
        
        # Export Input data
        sim.ExportInputData()    
        
        # Export Output data
        # sim.ExportOutputData(False, True)
        
        
    elif experimentType == 7: # UAV & MCS 갯수 증가
        
        # Set arm of discrete triangular distribution
        
        arm_Tri = 2
        
        # Generate State
        state = State()
        
        
        # Set the random seed for customer locations
        rndEVLocationGenerator = random.Random(24224) 
        
        # Set the random seed for customer demand
        rndEVDemandGenerator = random.Random(312222) 
        
        # Number of simulation
        numSims = 100
        
        
        
        # ground vehicle distnace weight
        groundVehicleDistanceWeight = 1.5
        
        # Save the results
        result_list = []
        
        for numSim in range(numSims):
        
            seed_EVsLocation = int(round(rndEVLocationGenerator.random() * 10000, 0))
            seed_EVsDemand = int(round(rndEVDemandGenerator.random() * 10000, 0))
            
            
            # Generate and assign EVs (customers)
            numEVs = 20
            evList_original = []
            GenerateEVs(numEVs, seed_EVsLocation, seed_EVsDemand, evList_original, groundVehicleDistanceWeight, arm_Tri)
            
            
            numEVList = [2, 4, 6, 8, 10, 12, 14, 16, 18, numEVs]
            numUAVsList = [1, 2, 3, 4, 5, 6]
            
            for numEVs in numEVList:
                for numUAVs in numUAVsList:     
                    evList = evList_original[0:numEVs] 
                    numMCSs = numUAVs
                
                
                    # EV를 UAV 별로 몇개씩 할당할지 결정 (UAV 당 최소 EV 1개씩은 할당시킴)
                    rndEVDivideGenerator = random.Random(seed_EVsLocation) 
                    
                    numAssignedEVs_UAV = []
                    cumNumEVs = 0
                    emptyUAVs = 0
                    for i in range(numUAVs):
                        if cumNumEVs == numEVs:
                            numAssignedEVs_UAV.append(0)
                            emptyUAVs += 1
                        else:
                            numAssignedEVs_UAV.append(1)
                            cumNumEVs += 1
                    
                    remainingNumEVs = numEVs - cumNumEVs
                        
                    for i in range(numUAVs):
                        if numAssignedEVs_UAV[i] != 0:
                            if i == numUAVs - emptyUAVs - 1:
                                numAssignedEVs_UAV[i] += remainingNumEVs
                                break
                            numAssignedEVs = rndEVDivideGenerator.randint(min(1, remainingNumEVs), remainingNumEVs)
                            remainingNumEVs -= numAssignedEVs
                            numAssignedEVs_UAV[i] += numAssignedEVs
                    
                
                   
                
                    # EV를 UAV 별로 할당하기 위한 list 준비
                    evList_UAVs = []
                    lastIndex = 0
                    cumNumAssignedEVs = 0 
                    for numAssignedEVs in numAssignedEVs_UAV:
                        tempEVList = []
                        tempEVList.extend(evList[lastIndex : lastIndex + numAssignedEVs])
                        evList_UAVs.append(tempEVList)
                        cumNumAssignedEVs += numAssignedEVs
                        lastIndex = cumNumAssignedEVs
                
                
                    # EV의 index를 UAV별로 0부터 시작하게 만들기 (OD matrix & energy matrix와 연관있음)
                    for evList_UAV in evList_UAVs:
                        for index, ev in enumerate(evList_UAV):
                            ev.index = index
                
                    uavList = []
                    mcsList = []
                    for i in range(numUAVs - emptyUAVs):
                        #### Generate an UAV
                        uav = UAV(cost_failure, cost_chargeFee, cost_flight, unitEnergyConsumption, uavCapacity, state, numEVs, chargingRate, startTime, uavSpeed, evList_UAVs[i], unitDiscrete)
                        # Set the depot
                        uav.setDepot(depotLoc[0], depotLoc[1])
                        
                        
                        # Generate OD matrices
                        uav.odMatrix = ODMatrix(uav.depot, uav.evList)
                        
                        # Generate Energy consumption matrices
                        uav.energyMatrix = EnergyMatrix(uav.energyConsumptionRate, uav.odMatrix, uav.evList, state)
                        
                        # Generate Cost matrices
                        uav.costMatrix = CostMatrix(uav.unitFlightCost, uav.energyMatrix, uav.evList)
                        
                        # Define DP
                        uav.dp = DynamicPrograming(uav, M)
                        
                        uavList.append(uav)
                    
                        #### Generate an MCS
                        mcs = MCS(mcsCapacity, state, chargingRate, startTime, evList_UAVs[i], unitDiscrete)
                        # Set the depot
                        mcs.setDepot(depotLoc[0], depotLoc[1])
                        
                        # Generate OD matrices
                        mcs.odMatrix = ODMatrix_MCS(mcs.depot, uav.evList, groundVehicleDistanceWeight)
                        
                        # Generate Energy consumption matrices
                        mcs.energyTimeMatrix = EnergyTimeMatrix_MCS(mcs.odMatrix, mcs.evList, state, seed_EVsDemand, UnitDistPowerConsumption_mean0, UnitDistPowerConsumption_mean1,
                                                            UnitDistPowerConsumption_mean2, UnitDistPowerConsumption_mean3, speed0, speed1, speed2, speed3, arm_Tri)
                        
                        
                        mcsList.append(mcs)
                        
                    
                        
                    # Sim
                    sim = Sim(M, uavList, mcsList, numEVs, result_list, len(uavList) + emptyUAVs, False, True, False, True)    
                    
                    # Export Input data
                    # sim.ExportInputData()    
                    
                    # UAV 초기화
                    # InitializeUAVs(uavList)
                    
                    # MCS 초기화
                    # InitializeMCSs(mcsList)
                # Export Output data
                # sim.ExportOutputData(False, True)
        
        
        # Export simulation result
        ExportSensitivityAnalysis_NumUAV_MCS_EV(result_list)
        
        # Export graphs
        sim.Graph_UAV_MCS(result_list)
        
        # Graph display
        print('test')
